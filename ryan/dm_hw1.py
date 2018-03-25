#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Created on Thu Mar 22 10:00:28 2018

@author: HSIN
"""

import sys
import time
import numpy as np
import openpyxl

from sklearn.utils import shuffle

from sklearn import svm
from sklearn.ensemble import RandomForestClassifier
from sklearn.ensemble import ExtraTreesClassifier
#from xgboost.sklearn import XGBClassifier

from sklearn.decomposition import PCA
from sklearn.decomposition import NMF

from sklearn.model_selection import cross_validate
from sklearn.model_selection import cross_val_score
from sklearn.model_selection import ShuffleSplit
from sklearn import metrics


stime = time.time()

np.random.seed(46)
eps = 1e-05

load_path = './super_market_data.xlsx'


wb = openpyxl.load_workbook(load_path)

ws = wb['交易記錄檔']

ws_max_row = ws.max_row
ws_max_col = ws.max_column


mids = set()
items = set()
item_classes = set()

for i in range(ws_max_row):
    
    the_row_idx = i + 1
    
    if the_row_idx == 1:
        continue
    
    mids.add(ws['B' + str(the_row_idx)].value)
    items.add(ws['F' + str(the_row_idx)].value)
    item_classes.add(ws['E' + str(the_row_idx)].value)


mids = sorted(list(mids))
items = sorted(list(items))
item_classes = sorted(list(item_classes))

x = np.zeros((len(mids),len(items)))
y = np.zeros(len(mids))


for i in range(ws_max_row):
    
    the_row_idx = i + 1
    
    if the_row_idx == 1:
        continue
    
    mid = ws['B' + str(the_row_idx)].value
    item = ws['F' + str(the_row_idx)].value
    item_class = ws['E' + str(the_row_idx)].value
    
    num = int(ws['G' + str(the_row_idx)].value)
    
    
    x[mids.index(mid), items.index(item)] += num
    
    

for i in range(x.shape[0]):
    for j in range(x.shape[1]):
        if x[i,j] < 0:
#            print(mids[i], items[j], x[i,j])
#            sys.exit()
            x[i,j] = 0




ws = wb['會員資料檔']

ws_max_row = ws.max_row
ws_max_col = ws.max_column


for i in range(ws_max_row):
    
    the_row_idx = i + 1
    
    if the_row_idx == 1:
        continue
    
    mid = ws['A' + str(the_row_idx)].value
    sex = ws['C' + str(the_row_idx)].value
    marry = ws['G' + str(the_row_idx)].value
    
    if sex == '男':
        y[mids.index(mid)] = 1
    elif sex == '女':
        y[mids.index(mid)] = 0
    else:
        print('Warning')
    
    
#    if marry == '已婚':
#        y[mids.index(mid)] = 1
#    else:
#        y[mids.index(mid)] = 0




#print(x.shape)
#print(y.shape)


# sugffle data
x, y = shuffle(x, y)



## Normalization
#x_mean = np.mean(x)
#x_std = np.std(x) + eps
#x = x - np.tile(x_mean,(len(x),1))
#x = x/np.tile(x_std,(len(x),1))


# modle building
svm = svm.LinearSVC()
rf = RandomForestClassifier(n_jobs = -1)
ext = ExtraTreesClassifier(n_jobs = -1)
#xgb = XGBClassifier(n_jobs = -1)



clfs = [svm, rf, ext]
dims = [x.shape[1],256,128,64,32,16,8,4,2]

scoring = ['accuracy','f1','precision','recall','roc_auc']



for dim in dims:  
    
    
    if dim == x.shape[1]:
        
        
        
        #################################################
        ################ NO Decomposition ###############
        #################################################
        
        print('-----------------------------------------------------')
        print('No Decomposition')
        
        print('x shape', x.shape)
        print('y shape', y.shape)
        
        for clf in clfs:
            
            print('-----------------------------------------------------')
            print('The Classifier')
            print(clf)
            print('-----------------------------------------------------')
            
            
        
            scores = cross_validate(clf,
                                    x, y,
                                    scoring = scoring,
                                    cv = 10,
                                    return_train_score = True,
                                    n_jobs = -1,
                                    verbose = 1
                                    )
            
            the_score_keys = sorted(scores.keys())
            
            for the_score_key in the_score_keys:
                the_scores = scores[the_score_key]
                print(the_score_key, ": %0.2f (+/- %0.2f)" % (the_scores.mean(), the_scores.std() * 2))
            
            print('-----------------------------------------------------')
    
    else:
        
            #################################################
            ################ PCA Decomposition ##############
            #################################################
        
            print('-----------------------------------------------------')
            print('PCA Decomposition', dim)
            
            pca = PCA(n_components = dim)
            x_pca = pca.fit_transform(x)
            y_pca = y
            
            
            print('x_pca shape', x_pca.shape)
            print('y_pca shape', y_pca.shape)
            
            for clf in clfs:
            
                print('-----------------------------------------------------')
                print('The Classifier')
                print(clf)
                print('-----------------------------------------------------')
                
                
            
                scores = cross_validate(clf,
                                        x_pca, y_pca,
                                        scoring = scoring,
                                        cv = 10,
                                        return_train_score = True,
                                        n_jobs = -1,
                                        verbose = 1
                                        )
                
                the_score_keys = sorted(scores.keys())
                
                for the_score_key in the_score_keys:
                    the_scores = scores[the_score_key]
                    print(the_score_key, ": %0.2f (+/- %0.2f)" % (the_scores.mean(), the_scores.std() * 2))
                
                print('-----------------------------------------------------')
            
            
            
            
            #################################################
            ################ NMF Decomposition ##############
            #################################################
            
            
            print('-----------------------------------------------------')
            print('NMF Decomposition', dim)
            
            nmf = NMF(n_components = dim)
            x_nmf = nmf.fit_transform(x)
            y_nmf = y
            
            
            print('x_nmf shape', x_nmf.shape)
            print('y_nmf shape', y_nmf.shape)
        
            
            
            
            for clf in clfs:
            
                print('-----------------------------------------------------')
                print('The Classifier')
                print(clf)
                print('-----------------------------------------------------')
                
                
                scores = cross_validate(clf,
                                            x_pca, y_pca,
                                            scoring = scoring,
                                            cv = 10,
                                            return_train_score = True,
                                            n_jobs = -1,
                                            verbose = 1
                                        )
                
                the_score_keys = sorted(scores.keys())
                
                for the_score_key in the_score_keys:
                    the_scores = scores[the_score_key]
                    print(the_score_key, ": %0.2f (+/- %0.2f)" % (the_scores.mean(), the_scores.std() * 2))
                
                print('-----------------------------------------------------')
    
    
    
    
    print('\n\n')




print('ALL Time Taken:', time.time()-stime)