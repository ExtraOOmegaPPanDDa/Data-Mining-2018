#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Created on Thu Mar 22 10:00:28 2018

@author: HSIN
"""

import sys
import csv
import time
import numpy as np
import matplotlib.pyplot as plt


import openpyxl

import math

from sklearn.utils import shuffle

from sklearn import svm
from sklearn.ensemble import RandomForestClassifier
from sklearn.ensemble import ExtraTreesClassifier
from xgboost.sklearn import XGBClassifier

from sklearn.decomposition import PCA
from sklearn.decomposition import NMF

from sklearn.model_selection import cross_validate
#from sklearn.model_selection import cross_val_score
#from sklearn.model_selection import ShuffleSplit
#from sklearn import metrics


stime = time.time()

np.random.seed(46)
eps = 1e-05
stime = time.time()



#################################################
# load data
#################################################

load_path = './super_market_data.xlsx'
predict_target = 'sex'

wb = openpyxl.load_workbook(load_path)

ws = wb['交易記錄檔']

ws_max_row = ws.max_row
ws_max_col = ws.max_column


mids = set()
items = set()
item_classes = set()

for i in range(ws_max_row):
    
    the_row_idx = i + 1
    
    if the_row_idx == 1:
        continue
    
    mids.add(ws['B' + str(the_row_idx)].value)
    items.add(ws['F' + str(the_row_idx)].value)
    item_classes.add(ws['E' + str(the_row_idx)].value)


mids = sorted(list(mids))
items = sorted(list(items))
item_classes = sorted(list(item_classes))

x = np.zeros((len(mids),len(items)))
y = np.zeros(len(mids)).astype('<U20')


for i in range(ws_max_row):
    
    the_row_idx = i + 1
    
    if the_row_idx == 1:
        continue
    
    mid = ws['B' + str(the_row_idx)].value
    item = ws['F' + str(the_row_idx)].value
    item_class = ws['E' + str(the_row_idx)].value
    
    num = int(ws['G' + str(the_row_idx)].value)
    
    
    x[mids.index(mid), items.index(item)] += num
    
    

for i in range(x.shape[0]):
    for j in range(x.shape[1]):
        if x[i,j] < 0:
#            print(mids[i], items[j], x[i,j])
#            sys.exit()
            x[i,j] = 0




ws = wb['會員資料檔']

ws_max_row = ws.max_row
ws_max_col = ws.max_column


for i in range(ws_max_row):
    
    the_row_idx = i + 1
    
    if the_row_idx == 1:
        continue
    
    mid = ws['A' + str(the_row_idx)].value
    
    sex = ws['C' + str(the_row_idx)].value
    marry = ws['G' + str(the_row_idx)].value
    income = ws['I' + str(the_row_idx)].value
    
    if predict_target == 'sex':
        y[mids.index(mid)] = sex
    elif predict_target == 'marry':
        y[mids.index(mid)] = marry
    elif predict_target == 'income':
        y[mids.index(mid)] = income
    else:
        print('Target not in the scope')
        sys.exit()



#################################################
# suffle data and balance-sample
#################################################

x, y = shuffle(x, y)


the_selected_idx = []
for i in range(len(y)):
    if y[i] != '其它':
        the_selected_idx.append(i)

x = x[the_selected_idx,:]
y = y[the_selected_idx]



print('x shape', x.shape)
print('y shape', y.shape)


print('Predict Target:', predict_target)
print('Category', set(y))

#sys.exit()



#################################################
# PCA Components
#################################################

pca = PCA(n_components = 2)
x_pca = pca.fit_transform(x)
plt.scatter(x_pca[:,0], x_pca[:,1], c=y)
#plt.show()
plt.savefig('market_pca_scatter.png')
plt.clf()



#################################################
# NMF Components
#################################################



nmf = NMF(n_components = 2)
x_nmf = nmf.fit_transform(x)
plt.scatter(x_nmf[:,0], x_nmf[:,1], c=y)
#plt.show()
plt.savefig('market_nmf_scatter.png')
plt.clf()




#################################################
# modle building
#################################################

svm = svm.LinearSVC()
rf = RandomForestClassifier(n_jobs = -1)
ext = ExtraTreesClassifier(n_jobs = -1)
xgb = XGBClassifier(n_jobs = -1)

clfs = [svm, rf, ext, xgb]
clf_names = ['SVM', 'RF', 'EXT', 'XGB']





#################################################
# scoring metrics
#################################################


#scoring = ['fit_time','accuracy','f1','precision','recall','roc_auc']
scoring = ['accuracy','f1_micro','f1_macro']


pca_fit_time_records = {}

pca_test_accuracy_records = {}
pca_test_f1_macro_records = {}
pca_test_f1_micro_records = {}

pca_train_accuracy_records = {}
pca_train_f1_macro_records = {}
pca_train_f1_micro_records = {}


nmf_fit_time_records = {}

nmf_test_accuracy_records = {}
nmf_test_f1_macro_records = {}
nmf_test_f1_micro_records = {}

nmf_train_accuracy_records = {}
nmf_train_f1_macro_records = {}
nmf_train_f1_micro_records = {}




for clf_name in clf_names:
    
    pca_fit_time_records[clf_name] = []

    pca_test_accuracy_records[clf_name] = []
    pca_test_f1_macro_records[clf_name] = []
    pca_test_f1_micro_records[clf_name] = []
    
    pca_train_accuracy_records[clf_name] = []
    pca_train_f1_macro_records[clf_name] = []
    pca_train_f1_micro_records[clf_name] = []
    
    
    nmf_fit_time_records[clf_name] = []

    nmf_test_accuracy_records[clf_name] = []
    nmf_test_f1_macro_records[clf_name] = []
    nmf_test_f1_micro_records[clf_name] = []
    
    nmf_train_accuracy_records[clf_name] = []
    nmf_train_f1_macro_records[clf_name] = []
    nmf_train_f1_micro_records[clf_name] = []
    


dims = [x.shape[1],512,256,128,64,32,16,8,4,2] 
#dims = [16,8,4,2]

pca_decomposition_times = []
nmf_decomposition_times = []



for dim in dims:  
    
    
    if dim == x.shape[1]:
        
        
        
        #################################################
        ################ NO Decomposition ###############
        #################################################
        
        print('-----------------------------------------------------')
        print('No Decomposition')
        
        print('x shape', x.shape)
        print('y shape', y.shape)
        
        for clf, clf_name in zip(clfs, clf_names):
            
            print('-----------------------------------------------------')
            print('The Classifier')
            print(clf)
            print('-----------------------------------------------------')
            
            
        
            scores = cross_validate(clf,
                                    x, y,
                                    scoring = scoring,
                                    cv = 10,
                                    return_train_score = True,
                                    n_jobs = -1,
                                    verbose = 1
                                    )
            
            the_score_keys = sorted(scores.keys())
            
            for the_score_key in the_score_keys:
                the_scores = scores[the_score_key]
                print(the_score_key, ": %0.2f (+/- %0.2f)" % (the_scores.mean(), the_scores.std() * 2))
            
            
            
            
            pca_fit_time_records[clf_name].append(scores['fit_time'].mean())
            pca_test_accuracy_records[clf_name].append(scores['test_accuracy'].mean())
            pca_test_f1_macro_records[clf_name].append(scores['test_f1_macro'].mean())
            pca_test_f1_micro_records[clf_name].append(scores['test_f1_micro'].mean())            
            pca_train_accuracy_records[clf_name].append(scores['train_accuracy'].mean())
            pca_train_f1_macro_records[clf_name].append(scores['train_f1_macro'].mean())
            pca_train_f1_micro_records[clf_name].append(scores['train_f1_micro'].mean())
            
            
            nmf_fit_time_records[clf_name].append(scores['fit_time'].mean())            
            nmf_test_accuracy_records[clf_name].append(scores['test_accuracy'].mean())
            nmf_test_f1_macro_records[clf_name].append(scores['test_f1_macro'].mean())
            nmf_test_f1_micro_records[clf_name].append(scores['test_f1_micro'].mean())            
            nmf_train_accuracy_records[clf_name].append(scores['train_accuracy'].mean())
            nmf_train_f1_macro_records[clf_name].append(scores['train_f1_macro'].mean())
            nmf_train_f1_micro_records[clf_name].append(scores['train_f1_micro'].mean())
            
            
            
            
            
            print('-----------------------------------------------------')
    
    else:
        
            #################################################
            ################ PCA Decomposition ##############
            #################################################
        
            print('-----------------------------------------------------')
            print('PCA Decomposition', dim)
            
            pca_stime = time.time()
            
            pca = PCA(n_components = dim)
            x_pca = pca.fit_transform(x)
            
            pca_decomposition_times.append(time.time() - pca_stime)
            
            y_pca = y
            
           
            
            
            print('x_pca shape', x_pca.shape)
            print('y_pca shape', y_pca.shape)
            
            for clf, clf_name in zip(clfs, clf_names):
            
                print('-----------------------------------------------------')
                print('The Classifier')
                print(clf)
                print('-----------------------------------------------------')
                
                
            
                scores = cross_validate(clf,
                                        x_pca, y_pca,
                                        scoring = scoring,
                                        cv = 10,
                                        return_train_score = True,
                                        n_jobs = -1,
                                        verbose = 1
                                        )
                
                the_score_keys = sorted(scores.keys())
                
                for the_score_key in the_score_keys:
                    the_scores = scores[the_score_key]
                    print(the_score_key, ": %0.2f (+/- %0.2f)" % (the_scores.mean(), the_scores.std() * 2))
                
               
                
                
                
                pca_fit_time_records[clf_name].append(scores['fit_time'].mean())
                pca_test_accuracy_records[clf_name].append(scores['test_accuracy'].mean())
                pca_test_f1_macro_records[clf_name].append(scores['test_f1_macro'].mean())
                pca_test_f1_micro_records[clf_name].append(scores['test_f1_micro'].mean())            
                pca_train_accuracy_records[clf_name].append(scores['train_accuracy'].mean())
                pca_train_f1_macro_records[clf_name].append(scores['train_f1_macro'].mean())
                pca_train_f1_micro_records[clf_name].append(scores['train_f1_micro'].mean())
                
                
                
                print('-----------------------------------------------------')
            
            
            
            
            #################################################
            ################ NMF Decomposition ##############
            #################################################
            
            
            print('-----------------------------------------------------')
            print('NMF Decomposition', dim)
            
            nmf_stime = time.time()
            
            nmf = NMF(n_components = dim)
            x_nmf = nmf.fit_transform(x)
            
            nmf_decomposition_times.append(time.time() - nmf_stime)
            
            y_nmf = y
            
            print('x_nmf shape', x_nmf.shape)
            print('y_nmf shape', y_nmf.shape)
        
            
            
            
            for clf, clf_name in zip(clfs, clf_names):
            
                print('-----------------------------------------------------')
                print('The Classifier')
                print(clf)
                print('-----------------------------------------------------')
                
                
                scores = cross_validate(clf,
                                            x_pca, y_pca,
                                            scoring = scoring,
                                            cv = 10,
                                            return_train_score = True,
                                            n_jobs = -1,
                                            verbose = 1
                                        )
                
                the_score_keys = sorted(scores.keys())
                
                for the_score_key in the_score_keys:
                    the_scores = scores[the_score_key]
                    print(the_score_key, ": %0.2f (+/- %0.2f)" % (the_scores.mean(), the_scores.std() * 2))
                
                
                
                nmf_fit_time_records[clf_name].append(scores['fit_time'].mean())            
                nmf_test_accuracy_records[clf_name].append(scores['test_accuracy'].mean())
                nmf_test_f1_macro_records[clf_name].append(scores['test_f1_macro'].mean())
                nmf_test_f1_micro_records[clf_name].append(scores['test_f1_micro'].mean())            
                nmf_train_accuracy_records[clf_name].append(scores['train_accuracy'].mean())
                nmf_train_f1_macro_records[clf_name].append(scores['train_f1_macro'].mean())
                nmf_train_f1_micro_records[clf_name].append(scores['train_f1_micro'].mean())
            
            
                
                
                print('-----------------------------------------------------')
    
    
    
    
    print('\n\n')


log_dims = [math.log(dim,2) for dim in dims]


#################################################
# fit time plot
#################################################

the_legends = []
for clf_name in clf_names:
    plt.plot(log_dims, pca_fit_time_records[clf_name])
    the_legends.append('PCA_' + clf_name)



plt.title('PCA Fit Time')
plt.ylabel('fit time')
plt.xlabel('log dim')
plt.legend(the_legends, loc='upper left')
#plt.show()
plt.savefig('market_pca_fit_time.png')
plt.clf()


the_legends = []
for clf_name in clf_names:
    plt.plot(log_dims, nmf_fit_time_records[clf_name])
    the_legends.append('NMF_' + clf_name)



plt.title('NMF Fit Time')
plt.ylabel('fit time')
plt.xlabel('log dim')
plt.legend(the_legends, loc='upper left')
#plt.show()
plt.savefig('market_nmf_fit_time.png')
plt.clf()



#################################################
# acc plot
#################################################

the_legends = []
for clf_name in clf_names:
    plt.plot(log_dims, pca_train_accuracy_records[clf_name])
    the_legends.append('PCA_Train_' + clf_name)

plt.title('PCA Train Accuracy')
plt.ylabel('train acc')
plt.xlabel('log dim')
plt.legend(the_legends, loc='auto')
plt.ylim(0,1.05)
#plt.show()
plt.savefig('market_pca_train_acc.png')
plt.clf()



the_legends = []
for clf_name in clf_names:
    plt.plot(log_dims, nmf_train_accuracy_records[clf_name])
    the_legends.append('NMF_Train_' + clf_name)

plt.title('NMF Train Accuracy')
plt.ylabel('train acc')
plt.xlabel('log dim')
plt.legend(the_legends, loc='auto')
plt.ylim(0,1.05)
#plt.show()
plt.savefig('market_nmf_train_acc.png')
plt.clf()



the_legends = []
for clf_name in clf_names:
    plt.plot(log_dims, pca_test_accuracy_records[clf_name])
    the_legends.append('PCA_Test_' + clf_name)

plt.title('PCA Test Accuracy')
plt.ylabel('test acc')
plt.xlabel('log dim')
plt.legend(the_legends, loc='auto')
plt.ylim(0,1.05)
#plt.show()
plt.savefig('market_pca_test_acc.png')
plt.clf()



the_legends = []
for clf_name in clf_names:
    plt.plot(log_dims, nmf_test_accuracy_records[clf_name])
    the_legends.append('NMF_Test_' + clf_name)

plt.title('NMF Test Accuracy')
plt.ylabel('test acc')
plt.xlabel('log dim')
plt.legend(the_legends, loc='auto')
plt.ylim(0,1.05)
#plt.show()
plt.savefig('market_nmf_test_acc.png')
plt.clf()



#################################################
# f1 macro plot
#################################################

the_legends = []
for clf_name in clf_names:
    plt.plot(log_dims, pca_train_f1_macro_records[clf_name])
    the_legends.append('PCA_Train_' + clf_name)

plt.title('PCA Train F1-Macro')
plt.ylabel('train f1-macro')
plt.xlabel('log dim')
plt.legend(the_legends, loc='auto')
plt.ylim(0,1.05)
#plt.show()
plt.savefig('market_pca_train_f1_macro.png')
plt.clf()



the_legends = []
for clf_name in clf_names:
    plt.plot(log_dims, nmf_train_f1_macro_records[clf_name])
    the_legends.append('NMF_Train_' + clf_name)

plt.title('NMF Train F1-Macro')
plt.ylabel('train f1-macro')
plt.xlabel('log dim')
plt.legend(the_legends, loc='auto')
plt.ylim(0,1.05)
#plt.show()
plt.savefig('market_nmf_train_f1_macro.png')
plt.clf()




the_legends = []
for clf_name in clf_names:
    plt.plot(log_dims, pca_test_f1_macro_records[clf_name])
    the_legends.append('PCA_Test_' + clf_name)

plt.title('PCA Test F1-Macro')
plt.ylabel('test f1-macro')
plt.xlabel('log dim')
plt.legend(the_legends, loc='auto')
plt.ylim(0,1.05)
#plt.show()
plt.savefig('market_pca_test_f1_macro.png')
plt.clf()



the_legends = []
for clf_name in clf_names:
    plt.plot(log_dims, nmf_test_f1_macro_records[clf_name])
    the_legends.append('NMF_Test_' + clf_name)

plt.title('NMF Test F1-Macro')
plt.ylabel('test f1-macro')
plt.xlabel('log dim')
plt.legend(the_legends, loc='auto')
plt.ylim(0,1.05)
#plt.show()
plt.savefig('market_nmf_test_f1_macro.png')
plt.clf()






#################################################
# f1 micro plot
#################################################

the_legends = []
for clf_name in clf_names:
    plt.plot(log_dims, pca_train_f1_micro_records[clf_name])
    the_legends.append('PCA_Train_' + clf_name)

plt.title('PCA Train F1-Micro')
plt.ylabel('train f1-micro')
plt.xlabel('log dim')
plt.legend(the_legends, loc='auto')
plt.ylim(0,1.05)
#plt.show()
plt.savefig('market_pca_train_f1_micro.png')
plt.clf()



the_legends = []
for clf_name in clf_names:
    plt.plot(log_dims, nmf_train_f1_micro_records[clf_name])
    the_legends.append('NMF_Train_' + clf_name)

plt.title('NMF Train F1-Micro')
plt.ylabel('train f1-micro')
plt.xlabel('log dim')
plt.legend(the_legends, loc='auto')
plt.ylim(0,1.05)
#plt.show()
plt.savefig('market_nmf_train_f1_micro.png')
plt.clf()




the_legends = []
for clf_name in clf_names:
    plt.plot(log_dims, pca_test_f1_micro_records[clf_name])
    the_legends.append('PCA_Test_' + clf_name)

plt.title('PCA Test F1-Micro')
plt.ylabel('test f1-micro')
plt.xlabel('log dim')
plt.legend(the_legends, loc='auto')
plt.ylim(0,1.05)
#plt.show()
plt.savefig('market_pca_test_f1_micro.png')
plt.clf()



the_legends = []
for clf_name in clf_names:
    plt.plot(log_dims, nmf_test_f1_micro_records[clf_name])
    the_legends.append('NMF_Test_' + clf_name)

plt.title('NMF Test F1-Micro')
plt.ylabel('test f1-micro')
plt.xlabel('log dim')
plt.legend(the_legends, loc='auto')
plt.ylim(0,1.05)
#plt.show()
plt.savefig('market_nmf_test_f1_micro.png')
plt.clf()





#################################################
# Decomposition Print
#################################################

print('Dims:', dims)
print('PCA Decomposition Time:', pca_decomposition_times)
print('NMF Decomposition Time:', nmf_decomposition_times)



print('All Time Taken:', time.time()-stime)