#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Created on Wed May  9 16:26:09 2018

@author: HSIN
"""



import time
import numpy as np
import openpyxl
import string
import pickle

from collections import Counter

import pyfpgrowth
#from pymining import itemmining, assocrules, perftesting

import openpyxl

import sys

import jieba

stime = time.time()


filter_words = set()

f = open('filter_words.txt')
for line in f:
    filter_words.add(line.rstrip())
f.close()

punct_chs = string.punctuation + "，；。、：「」『』《》（）【】！？％1234567890０１２３４５６７８９＊－／．"
translator = str.maketrans(punct_chs, ' ' * len(punct_chs))


def term_valid_check(ustr):
    
    for uchar in ustr:
        if ' ' == uchar or \
        uchar.isdigit():
            
            return False
    
    return True


def short_eng_check(term):
    
    short_check = True
    eng_check = True
    
    if len(term) > 2:
        short_check = False
    
    for uchar in term:
        if not ((u'\u0041' <= uchar <= u'\u005a') or (u'\u0061' <= uchar <= u'\u007a')):
            eng_check = False
    
    if short_check and eng_check:
        return True
    else:
        return False
            


load_path = './dataset.xlsx'

wb = openpyxl.load_workbook(load_path)
ws = wb['all']

ws_max_row = ws.max_row
ws_max_col = ws.max_column

n_grams_n = [2,3,4,5,6,7]


keywords = ['鴻海', '富士康', '蘋果', 'Apple', '台積電', '中華電信', '選舉', '大數據']


stockwords = ['台股','個股','類股','收盤','上漲','下跌','漲幅']

N_docs = 0

tf_counter = Counter()
df_counter = Counter()

the_row_idx2the_tf_counter = {}



memory_release_step = 1500
tf_thresh_min = 1
tf_thresh_max = 30

to_run = ws_max_row

for i in range(to_run):
    
    the_row_idx = i + 1
    
    if the_row_idx == 1:
        continue
    
    
    if the_row_idx % memory_release_step == 0:
            
        print('\n')
        print('----------------------------------')
        print('Time Taken:', time.time() - stime)
        print('Release Memory...')
        
        origin_size = len(df_counter)
        
        tf_thresh = int(tf_thresh_min + ((tf_thresh_max - tf_thresh_min)/to_run) * the_row_idx)
        print('TF Thresh:', tf_thresh)
        
        tf_counter = Counter({k:tf_counter[k] for k in tf_counter if tf_counter[k] > tf_thresh})
        df_counter = tf_counter & df_counter
                   
        adjusted_size = len(df_counter)
        
        
        print('Origin Size', origin_size)
        print('Adjusted Size', adjusted_size)
        
        
        print('----------------------------------')
        print('\n')
    
    
    the_title = ws['D' + str(the_row_idx)].value
    the_content = ws['E' + str(the_row_idx)].value
    
    stock_check = False
    
    for stockword in stockwords:
        if stockword in the_title or stockword in the_content:
            stock_check = True
            break
    
    if stock_check:
        continue
    
    if '中央氣象局' in the_title:
        continue
    
    the_doc_text = the_title + ' ' + the_content  
    the_doc_text = the_doc_text.translate(translator)
    
    
    selected = False
    
    for keyword in keywords:
        if keyword in the_doc_text:
            selected = True
            break
    
    if selected:
        N_docs += 1
    else:
        continue
        
    
    the_tf_counter = Counter()
    
    for n in n_grams_n:
        for i in range(len(the_doc_text)-n+1):
            term = the_doc_text[i:i+n]
            if term_valid_check(term):
                the_tf_counter[term] += 1
                
    for term in the_tf_counter:
        tf_counter[term] += the_tf_counter[term]
        df_counter[term] += 1
    
    
    the_row_idx2the_tf_counter[the_row_idx] = the_tf_counter



print('\n')
print('----------------------------------')
print('Time Taken:', time.time() - stime)
print('Release Memory...')

origin_size = len(tf_counter)

tf_thresh = tf_thresh_max

tf_counter = tf_counter = Counter({k:tf_counter[k] for k in tf_counter if tf_counter[k] > tf_thresh})
df_counter = tf_counter & df_counter

adjusted_size = len(tf_counter)
        
        
print('Origin Size', origin_size)
print('Adjusted Size', adjusted_size)

print('----------------------------------')
print('\n')


print('\n')
print('----------------------------------')
print('Time Taken:', time.time() - stime)
print('SE Score Filtering...')

origin_size = len(tf_counter)

term_list = sorted(tf_counter, key = len, reverse = True)

del_term_set = set()

se_thresh1 = 0.6
se_thresh2 = 0.6

eps = 1e-5


for term in term_list:
    
    if short_eng_check(term):
        del_term_set.add(term)
    
    if len(term) == max(n_grams_n):
        del_term_set.add(term)
    
    if len(term) > 2:
            
        c = term
        a = term[:-1]
        b = term[1:]
        
        se_score = tf_counter[c]/(tf_counter[a] + tf_counter[b] - tf_counter[c] + eps)
        
        if se_score > se_thresh1:
            
            del_term_set.add(a)
            del_term_set.add(b)
        
        if tf_counter[c]/tf_counter[a] > se_thresh2:
            del_term_set.add(a)
        
        if tf_counter[c]/tf_counter[b] > se_thresh2:
            del_term_set.add(b)
        

for term in del_term_set:
    del tf_counter[term]

adjusted_size = len(tf_counter)

df_counter = tf_counter & df_counter

print('Origin Size', origin_size)
print('Adjusted Size', adjusted_size)

print('----------------------------------')
print('\n')



for term in tf_counter:
    jieba.add_word(term, tf_counter[term] * 100000)



the_term_sets = []


for i in range(ws_max_row):
    
    the_row_idx = i + 1
    
    if the_row_idx == 1:
        continue
    
    the_title = ws['D' + str(the_row_idx)].value
    the_content = ws['E' + str(the_row_idx)].value
    
    stock_check = False
    
    for stockword in stockwords:
        if stockword in the_title or stockword in the_content:
            stock_check = True
            break
    
    if stock_check:
        continue
    
    if '中央氣象局' in the_title:
        continue
    
    
    the_doc_text = ws['D' + str(the_row_idx)].value + ' ' + ws['E' + str(the_row_idx)].value     
    the_doc_text = the_doc_text.translate(translator)
    
    selected = False
    
    for keyword in keywords:
        if keyword in the_doc_text:
            selected = True
            break
    
    if not selected:
        continue
    
    
    
    
    the_seg_list = jieba.cut(the_doc_text)
    
    the_seg_list = list(filter(lambda a: a != ' ', the_seg_list))
    
    the_term_set = set()
    
    for the_term in the_seg_list:
        if the_term in tf_counter and the_term not in filter_words:
            the_term_set.add(the_term)
            
    
    the_term_sets.append(list(the_term_set))
    


print('frequent pattern mining...')

the_input = the_term_sets[:]
sup_min = 50
conf_min = 0.6

print('count', len(the_input))
print('sup_min:', sup_min)
print('conf_min:', conf_min)

patterns = pyfpgrowth.find_frequent_patterns(the_input, sup_min)
print('PATTERN')

nodes = set()
for the_key in sorted(patterns, key = len, reverse = True)[:1000]:
    the_result = patterns[the_key]
    print(the_key, the_result)
    
    for item in the_key:
        nodes.add(item)

nodes = sorted(list(nodes))

adjacency_mat = np.zeros((len(nodes),len(nodes)))

for the_key in sorted(patterns, key = len, reverse = True)[:1000]:
    the_result = patterns[the_key]
#    print(the_key, the_result)
    for item in the_key:
        for item2 in the_key:
            adjacency_mat[nodes.index(item), nodes.index(item2)] += the_result


for i in range(len(adjacency_mat)):
    adjacency_mat[i,:] /= sum(adjacency_mat[i,:])
            
        

print('\n\n')


rules = pyfpgrowth.generate_association_rules(patterns, conf_min)
print('RULE')
for the_key in sorted(rules, key = len, reverse = True)[:1000]:
    the_result = rules[the_key]
    print(the_key, the_result)

#relim_input = itemmining.get_relim_input(the_input)
#report = itemmining.relim(relim_input, min_support = 20)




import matplotlib.pyplot as plt
from sklearn.manifold import TSNE
#from sklearn.decomposition import PCA
#from sklearn.decomposition import NMF
from sklearn.cluster import KMeans
from adjustText import adjust_text

plt.rcParams['font.sans-serif']=['SimHei'] 
plt.rcParams['axes.unicode_minus'] = False


kmeans_result = KMeans(n_clusters = 10, random_state = 0).fit(adjacency_mat)


decomposition_model = TSNE(n_components = 2)
np.set_printoptions(suppress = True)
vis_data = decomposition_model.fit_transform(adjacency_mat) 

vis_x = vis_data[:,0]
vis_y = vis_data[:,1]


#plt.figure(figsize=(16, 9))
#plt.scatter(vis_x, vis_y, c = kmeans_result.labels_)
#for label, x, y in zip(nodes, vis_x, vis_y):
#    plt.annotate(label, xy=(x, y), xytext=(0, 0), textcoords='offset points')
#
#
#
#plt.savefig('draw.png', dpi = 1000)
#plt.show()



def plot_scatter(adjust, xvalue, yvalue, label, color):
    plt.clf()
    plt.figure(figsize=(16, 9))
    plt.scatter(xvalue, yvalue, s = 15, c = color, edgecolors = 'None', alpha = 0.5)
    texts = []
    for x, y, s in zip(xvalue, yvalue, label):
        texts.append(plt.text(x, y, s, size=7))
    if adjust:
        adjust_text(texts, arrowprops = dict(arrowstyle = "-", color = 'k', lw = 0.5))
    
    
    plt.savefig('draw.png', dpi = 1000)
    plt.show()
        
plot_scatter(adjust = True, xvalue = vis_x, yvalue = vis_y, label = nodes, color = kmeans_result.labels_)

